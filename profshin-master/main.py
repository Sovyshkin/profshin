from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
from flask_cors import CORS, cross_origin

import os.path
import json
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
import traceback
import os
import telebot
from telebot import types
import threading
import contextlib
from datetime import datetime
import locale
#from flask_migrate import Migrate


BASE_PATH = os.path.dirname(os.path.realpath(__file__))

# connection = sqlite3.connect("db.sqlite3", check_same_thread=False)
# cursor = connection.cursor()


#app = Flask(__name__)
app = Flask(__name__, static_folder='./dist', static_url_path='')
# app.config['CORS_HEADERS'] = 'Content-Type'

cors = CORS(app, resources={r"/*": {"origins": "*"}})
#migrate = Migrate(app, db)
#version = "v1"
#cors = CORS(app)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///price.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app) 
app.app_context().push()

class Price(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    types_data = db.Column(db.Text, nullable=True) # вид
    kod_data = db.Column(db.Text, nullable=True) # код
    tmc_data = db.Column(db.Text, nullable=True) # тмц
    diameter_data = db.Column(db.Text, nullable=True) # дифметр
    width_data = db.Column(db.Text, nullable=True) # ширина
    loosening_data = db.Column(db.Text, nullable=True) # разболтовка
    hub_data = db.Column(db.Text, nullable=True) # ступица
    departure_data = db.Column(db.Text, nullable=True) # вылет
    brends_data = db.Column(db.Text, nullable=True) # бренд
    price_data = db.Column(db.Text, nullable=True) # сайт цена
    href_data = db.Column(db.Text, default='./src/assets/img/no_photo.jpg') # ссылка на картинку

class Chat(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ticket = db.Column(db.Integer, nullable=True)
    message = db.Column(db.Text, nullable=True)
    time = db.Column(db.Text, nullable=True)
    ans = db.Column(db.Boolean, nullable=True)

    def to_dict(self):
        return {
            'id': self.id,
            'ticket': self.ticket,
            'message': self.message,
            'time': self.time,
            'ans': self.ans

        }

db.create_all()


def start_bot():
    TOKEN = "6490087475:AAEzhup6IQDhP5ko8aftyxvrp6zvDrgncEs"
    bot = telebot.TeleBot(TOKEN)

    @bot.message_handler(commands=["start"])
    def start(message):
        if str(message.chat.id) == '1363777899' or str(message.chat.id) == '4142509878': 
            bot.send_message(message.chat.id, "Hello")

    def save():
        try:
            os.remove(f"./instance/price.db")
        except:
            pass
        with app.app_context():
            db.create_all()
        print("ok")
        wb = load_workbook('./src/assets/price/pricelist_gruz.xlsx')
        sheet = wb['Worksheet']
        counter = 1
        while sheet[f'A{counter}'].value != None:
            types = sheet[f'A{counter}'].value 
            kod = sheet[f'B{counter}'].value 
            tmc = sheet[f'C{counter}'].value 
            diameter = sheet[f'D{counter}'].value             
            width = sheet[f'E{counter}'].value
            loosening = sheet[f'F{counter}'].value
            hub = sheet[f'G{counter}'].value
            departure = sheet[f'H{counter}'].value
            brends = sheet[f'I{counter}'].value
            price = sheet[f'J{counter}'].value
            if sheet[f'N{counter}'].value != None:
                href = sheet[f'N{counter}'].value
                prices = Price(types_data = types, # вид
                        kod_data = kod, # код
                            tmc_data = tmc, # тмц
                            diameter_data = diameter,# дифметр
                            width_data = width, # ширина
                            loosening_data = loosening, # разболтовка
                            hub_data = hub, # ступица
                            departure_data = departure, # вылет
                            brends_data = brends,  # бренд
                            price_data = price, # сайт цена
                            href_data = href)# ссылка на картинку
            else:          
                prices = Price(types_data = types, # вид
                            kod_data = kod, # код
                                tmc_data = tmc, # тмц
                                diameter_data = diameter,# дифметр
                                width_data = width, # ширина
                                loosening_data = loosening, # разболтовка
                                hub_data = hub, # ступица
                                departure_data = departure, # вылет
                                brends_data = brends,  # бренд
                                price_data = price) # сайт цена 
            try:
                with app.app_context():
                    db.session.add(prices)
                    db.session.commit()
                counter += 1
            except:
                a = "Error DataBase"
                traceback.print_exc()
                print(a)
                return a
        wb.close()
        #os.remove(f"./src/assets/price/pricelist_gruz.xlsx")

    @bot.message_handler(content_types=["document"])
    def handle_docs(message):
        try:
            os.remove(f"./src/assets/price/pricelist_gruz.xlsx")
        except:
            pass
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        src = './src/assets/price/' + "pricelist_gruz.xlsx"#message.document.file_name
        with open(src, 'wb') as new_file:
            new_file.write(downloaded_file)
        save()

    bot.infinity_polling(timeout=50, long_polling_timeout=50)
bot_thread = threading.Thread(target=start_bot)
bot_thread.start()


def start_bot_two():

    TOKEN = "7075413939:AAEJupsb3LfsJuCcpSku9l7ZNf5iXfL74TM"
    bot = telebot.TeleBot(TOKEN)

    @bot.message_handler(commands=["start"])
    def start(message):
        if str(message.chat.id) == '1613828142': 
            bot.send_message(1613828142, "Привет, я готов принимать сообщения")
    @bot.message_handler(commands=["ticket"])
    def start(message):
        if str(message.chat.id) == '1613828142':
            text = str(message.text).split()
            num_ticket = int(text[2][:-1])
            msg = ' '.join(text[3:])
            current_datetime = datetime.now()
            time = current_datetime.strftime("%d %B, %H:%M")
            with app.app_context():
                chat_msg = Chat(message=msg, ticket=num_ticket, time=time, ans=True)
                db.session.add(chat_msg)
                db.session.commit()
            bot.send_message(1613828142, 'Отправлено...')
            
    @bot.message_handler(func=lambda message: True)
    def echo_all(message):
        bot.reply_to(message, message.text)


    bot.infinity_polling(timeout=50, long_polling_timeout=50)

bot_thread_two = threading.Thread(target=start_bot_two)
bot_thread_two.start()



@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/src/assets/img/no_photo.jpg')
def img():
    filename = './src/assets/img/no_photo.jpg'
    return send_file(filename, mimetype='image/gif')

@app.route('/catalog/src/assets/img/no_photo.jpg')
def img2():
    filename = './src/assets/img/no_photo.jpg'
    return send_file(filename, mimetype='image/gif')

@app.route('/filter', methods=['POST', 'GET'])#f'/{version}/filter'
#@cross_origin()
def filter() -> json:
    try:
        result = {}
        if request.method == "POST":
            # typeCar = request.form['typeCar']
            # count_tr = request.form['count_tr']
            # typeSensor = request.form['typeSensor']
            # integrationMon = request.form['integrationMon']
            wb = load_workbook('./src/assets/price/pricelist_gruz.xlsx')
            sheet = wb['Worksheet']
            counter = 1
                
    #------------------#
    #--промежуточные--#
    #------------------#
            diameter_pDisk = []
            width_pDisk = []
            loosening_pDisk = []
            hub_pDisk = []
            departure_pDisk = []
            brends_pDisk = []
            price_pDisk = []
            price_pTires =[]
            brend_pTires = []
            tmc_pTires = []
    #------------------#

    #------------------#
    #--в фильтры--#
    #------------------#            
            diameterCargoDisk ={}
            widthCargoDisk ={}
            looseningCargoDisk ={}
            hubCargoDisk ={}
            departureCargoDisk ={}
            brendCargoDisk = {}
            priceCargoDisk = {}
            price_CargoTires ={}
            brend_CargoTires = {}
            tmc_CargoTires = {}
    #------------------#
            typeCar = "cargo"
            if typeCar == "cargo": #грузовой  
                while sheet[f'A{counter}'].value != None:
                    #print(sheet[f'A{counter}'].value)
                    if sheet[f'A{counter}'].value == "Грузовые диски":
                        print(sheet[f'I{counter}'].value)
                        diameter_pDisk.append(sheet[f'D{counter}'].value)              
                        width_pDisk.append(sheet[f'E{counter}'].value)
                        loosening_pDisk.append(sheet[f'F{counter}'].value)
                        hub_pDisk.append(sheet[f'G{counter}'].value)
                        departure_pDisk.append(sheet[f'H{counter}'].value)
                        brends_pDisk.append(sheet[f'I{counter}'].value)
                        price_pDisk.append(sheet[f'J{counter}'].value)

                    if sheet[f'A{counter}'].value == "Шины грузовые":
                        price_pTires.append(sheet[f'J{counter}'].value)
                        brend_pTires.append(sheet[f'I{counter}'].value)
                        tmc_pTires.append(sheet[f'C{counter}'].value)
                    counter += 1
                
                diameterCargoDisk = set(diameter_pDisk)
                widthCargoDisk = set(width_pDisk)
                looseningCargoDisk = set(loosening_pDisk)
                hubCargoDisk = set(hub_pDisk)
                departureCargoDisk = set(departure_pDisk)
                brendCargoDisk = set(brends_pDisk)
                priceCargoDisk = set(price_pDisk)

                price_CargoTires = set(price_pTires)
                brend_CargoTires = set(brend_pTires)
                tmc_CargoTires = set(tmc_pTires)

                result['diameterCargoDisk'] = list(diameterCargoDisk) 
                result['widthCargoDisk'] = list(widthCargoDisk) 
                result['looseningCargoDisk'] = list(looseningCargoDisk) 
                result['hubCargoDisk'] = list(hubCargoDisk) 
                result['departureCargoDisk'] = list(departureCargoDisk) 
                result['brendCargoDisk'] = list(brendCargoDisk) 
                result['priceCargoDiskk'] = list(priceCargoDisk) 
                result['price_CargoTires'] = list(price_CargoTires) 
                result['brend_CargoTires'] = list(brend_CargoTires)
                result['tmc_CargoTires'] = list(tmc_CargoTires)

                diameter_p = []
                width_p = []
                loosening_p = []
                hub_p = []
                departure_p = []
                brends_p = []
                price_p = []
                price_pTires =[]
                brend_pTires = []
                tmc_pTires = []

                wb.close()
            
            elif typeCar == "passenger": #легковой 
                while sheet[f'A{counter}'].value != None:
                    #print(sheet[f'A{counter}'].value)
                    if sheet[f'A{counter}'].value == "Легкогрузовые диски":
                        print(sheet[f'I{counter}'].value)
                        diameter_pDisk.append(sheet[f'D{counter}'].value)              
                        width_pDisk.append(sheet[f'E{counter}'].value)
                        loosening_pDisk.append(sheet[f'F{counter}'].value)
                        hub_pDisk.append(sheet[f'G{counter}'].value)
                        departure_pDisk.append(sheet[f'H{counter}'].value)
                        brends_pDisk.append(sheet[f'I{counter}'].value)
                        price_pDisk.append(sheet[f'J{counter}'].value)

                    if sheet[f'A{counter}'].value == "Шины грузовые":
                        price_pTires.append(sheet[f'J{counter}'].value)
                    counter += 1
                
                diameterCargoDisk = set(diameter_pDisk)
                widthCargoDisk = set(width_pDisk)
                looseningCargoDisk = set(loosening_pDisk)
                hubCargoDisk = set(hub_pDisk)
                departureCargoDisk = set(departure_pDisk)
                brendCargoDisk = set(brends_pDisk)
                priceCargoDisk = set(price_pDisk)

                price_CargoTires = set(price_pTires)
                brend_CargoTires = set(brend_pTires)

                result['diameterCargoDisk'] = list(diameterCargoDisk) 
                result['widthCargoDisk'] = list(widthCargoDisk) 
                result['looseningCargoDisk'] = list(looseningCargoDisk) 
                result['hubCargoDisk'] = list(hubCargoDisk) 
                result['departureCargoDisk'] = list(departureCargoDisk) 
                result['brendCargoDisk'] = list(brendCargoDisk) 
                result['priceCargoDiskk'] = list(priceCargoDisk) 
                result['price_CargoTires'] = list(price_CargoTires)
                result['brend_CargoTires'] = list(brend_CargoTires)

                diameter_p = []
                width_p = []
                loosening_p = []
                hub_p = []
                departure_p = []
                brends_p = []
                price_p = []
                price_pTires =[]
                brend_pTires = []

                wb.close()

            return jsonify(result)
        
    except Exception as e:
        print("Error: ", e)
        result = {}
        result['error'] = e
        return jsonify(result)



@app.route('/catalog', methods=['POST', 'GET'])#f'/{version}/catalog'
def catalog() -> json:
    try:
        print(request.headers)
        result = {}
        data = request.get_json()
        print(data)
        typeCar = data['params']['typeCar']
        data_filter = []
        catalogs = Price.query.order_by(Price.types_data).all()

        if str(data['params']['tires']) == "True":
            pass

        if str(data['params']['disk']) == "True":
            
            diameterCargoDisk = str(data['params']['diameterCargoDisk'])
            widthCargoDisk = str(data['params']['widthCargoDisk'])
            looseningCargoDiskn = str(data['params']['looseningCargoDisk'])
            hubCargoDisk = str(data['params']['hubCargoDisk'])
            departureCargoDisk = str(data['params']['departureCargoDisk'])
            brendCargoDisk = str(data['params']['brendCargoDisk'])
            try:
                getcatalog = str(data['params']['getcatalog'])
            except:
                getcatalog = True
            id = 0
            if getcatalog == False:
                for el in catalogs:
                    data_set = '{ "code": ' + f'"{el.kod_data}", "tmc": "{el.tmc_data}", "price": "{el.price_data}", "href": "{el.href_data}", "id": "{id}"' + ' }'
                    data_filter.append(data_set)
            else:
                for el in catalogs:
                    if typeCar == "cargo":
                        if str(el.types_data) == "Грузовые диски" or str(el.types_data) == "Шины грузовые":
                            if diameterCargoDisk == str(el.diameter_data) or diameterCargoDisk == "None":
                                if widthCargoDisk == str(el.width_data) or widthCargoDisk == "None":
                                    print(3)
                                    if looseningCargoDiskn == str(el.loosening_data) or looseningCargoDiskn == "None":
                                        print(4)
                                        if hubCargoDisk == str(el.hub_data) or hubCargoDisk == "None":
                                            print(5)
                                            if departureCargoDisk == str(el.departure_data) or departureCargoDisk == "None":
                                                print(6)
                                                if brendCargoDisk == str(el.brends_data) or brendCargoDisk == "None":
                                                    data_set = '{ "code": ' + f'"{el.kod_data}", "tmc": "{el.tmc_data}", "price": "{el.price_data}", "href": "{el.href_data}", "id": "{id}"' + ' }'
                                                    id += 1
                                                    print(data_set)
                                                    data_filter.append(data_set)
                    if typeCar == "passenger":
                        if str(el.types_data) == "Легкогрузовые диски" or str(el.types_data) == "Литые диски" or str(el.types_data) == "Штампованные диски":
                            if diameterCargoDisk == str(el.diameter_data) or diameterCargoDisk == "None":
                                if widthCargoDisk == str(el.width_data) or widthCargoDisk == "None":
                                    print(3)
                                    if looseningCargoDiskn == str(el.loosening_data) or looseningCargoDiskn == "None":
                                        print(4)
                                        if hubCargoDisk == str(el.hub_data) or hubCargoDisk == "None":
                                            print(5)
                                            if departureCargoDisk == str(el.departure_data) or departureCargoDisk == "None":
                                                print(6)
                                                if brendCargoDisk == str(el.brends_data) or brendCargoDisk == "None":
                                                    data_set = '{ "code": ' + f'"{el.kod_data}", "tmc": "{el.tmc_data}", "price": "{el.price_data}", "href": "{el.href_data}", "id": "{id}"' + ' }'
                                                    id += 1
                                                    print(data_set)
                                                    data_filter.append(data_set)

        print(data_filter)
        result['DataFilters'] = data_filter
        return jsonify(result)
    
    except Exception as e:
        print("Error: ", e)
        result = {}
        result['error'] = e
        return jsonify(result)


@app.route('/getcatalog', methods=['POST', 'GET'])#f'/{version}/getcatalog'
def getcatalog() -> json:
    try:
        result = {}
        data = request.get_json()
        print(data)
        data_filter = []
        catalogs = Price.query.order_by(Price.types_data).all()
        for el in catalogs:
            if f"{str(el.types_data)}" == "Грузовые диски" or f"{str(el.types_data)}" == "Шины грузовые":
                data_set = f'{{ "code": "{el.kod_data}", "tmc": "{el.tmc_data}", "price": "{el.price_data}", "href": "{el.href_data}", "cargo": "{str(el.types_data) == "Грузовые диски"}" }}'
                data_filter.append(data_set)
            if f"{str(el.types_data)}" == "Легкогрузовые диски" or f"{str(el.types_data)}" == "Литые диски" or f"{str(el.types_data)}" == "Штампованные диски":
                data_set = f'{{ "code": "{el.kod_data}", "tmc": "{el.tmc_data}", "price": "{el.price_data}", "href": "{el.href_data}", "cargo": "{str(el.types_data) == "Грузовые диски"}" }}'
                data_filter.append(data_set)

        result['DataFilters'] = data_filter
        return jsonify(result)
    
    except Exception as e:
        print("Error: ", e)
        result = {}
        result['error'] = e
        return jsonify(result)


@app.route('/set_card', methods=['POST', 'GET'])#f'/{version}/set_card'
def set_card() -> json:
    try:
        result = {}
        data_filter = []
        data = request.get_json()
        print(data)
        kod_get = str(data['params']['kod'])

        catalogs = Price.query.order_by(Price.types_data).all 
        for el in catalogs:
            if kod_get == str(el.kod_data):
                data_set = '{ "code": ' + f'"{el.kod_data}", "tmc": "{el.tmc_data}", "price": "{el.price_data}", "href": "{el.href_data}", "id": "{id}"' + ' }'
                print(data_set)
                data_filter.append(data_set)
                
        print(data_filter)
        result['DataFilters'] = data_filter
        return jsonify(result)
    except Exception as e:
        print("Error: ", e)
        result = {}
        result['error'] = e
        return jsonify(result)
    
@app.route(f'/send_message', methods=['POST', 'GET'])
def send_message() -> json:

    result = {}
    data = request.get_json()
    message = data['params']['message']
    ticket = data['params']['ticket']
    time = data['params']['time']
    print(ticket)
    if message and time:
        if ticket != 'False':
            msg = Chat(message=message, time=time, ticket=ticket, ans=False)
            db.session.add(msg)
            db.session.commit()
            result['status'] = '200'
        else:
            msg = Chat(message=message, time=time, ans=False)
            db.session.add(msg)
            db.session.commit()
            result['ticket'] = msg.id
            result['status'] = '200'
            msg.ticket = msg.id
            db.session.commit()

        TOKEN = "7075413939:AAEJupsb3LfsJuCcpSku9l7ZNf5iXfL74TM"
        bot = telebot.TeleBot(TOKEN)
        bot.send_message(1613828142, f'Номер тикета: {msg.ticket}. Сообщение: {message}')

    return jsonify(result)


@app.teardown_appcontext
def shutdown_session(exception=None):
    db.session.remove()

@contextlib.contextmanager
def session_scope():
    """Управляет сессией базы данных."""
    session = db.session
    try:
        yield session
    finally:
        session.close()


@app.route(f'/get_messages', methods=['POST', 'GET'])
def get_messages() -> json:
    data = request.get_json()
    ticket = data['params']['ticket']
    with session_scope() as session:
        messages = session.query(Chat).filter_by(ticket=ticket).all()
        messages = [chat.to_dict() for chat in messages]

    return jsonify(messages)



#if __name__ == "__main__":
    #app.run(host="127.0.0.1", debug=False)# port=5010
    #app.run(port=8080)
    #app.run(host="0.0.0.0", debug=False)# port=5010,
    #app.run(host="0.0.0.0", port=8080, debug=False)# port=5010